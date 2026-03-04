from pathlib import Path
from datetime import date, datetime
from typing import Any, Optional, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from cnv_etl.models.document import CleanFinancialStatement
from cnv_etl.models.company import Company
from cnv_etl.errors import PipelineReport
from cnv_etl.logging_config import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Export financial statements and pipeline reports to Excel format."""

    HEADER_FILL      = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    HEADER_FONT      = Font(bold=True, size=11)
    ERROR_FILL       = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    BORDER           = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        self.wb = Workbook()

    # ------------------------------------------------------------------ #
    # Company export                                                       #
    # ------------------------------------------------------------------ #

    def export_company(self, company: Company, output_path: Path) -> Path:
        if not company.statements:
            raise ValueError("Cannot export empty statements list")

        if output_path.suffix != '.xlsx':
            output_path = output_path.with_suffix('.xlsx')

        self.wb.remove(self.wb.active) # type: ignore

        self._write_company_metadata_sheet(company)
        self._write_statement_metadata_sheet(company)
        self._write_statement_values_sheet(company.statements)

        self.wb.save(output_path)
        logger.info(f"Exported {len(company.statements)} statements to {output_path}")

        return output_path

    # ------------------------------------------------------------------ #
    # Pipeline report export                                               #
    # ------------------------------------------------------------------ #

    def write_pipeline_report(self, report: PipelineReport, output_path: Path) -> Path:
        """
        Write a PipelineReport to an Excel file with two sheets:
          - run_summary : aggregate counts for the full run
          - errors      : one row per ETLError across all companies
        """
        if output_path.suffix != '.xlsx':
            output_path = output_path.with_suffix('.xlsx')

        self.wb.remove(self.wb.active) # type: ignore

        self._write_run_summary_sheet(report)
        self._write_errors_sheet(report)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        logger.info(f"Pipeline report saved to {output_path}")

        return output_path

    def _write_run_summary_sheet(self, report: PipelineReport) -> None:
        ws = self.wb.create_sheet("run_summary")

        rows = [
            ("started_at",                   report.started_at),
            ("duration_seconds",             round(report.duration_seconds, 2)),
            ("companies_attempted",          report.total_companies_attempted),
            ("companies_succeeded",          report.total_companies_succeeded),
            ("companies_failed",             report.total_companies_failed),
            ("statements_downloaded",        report.total_statements_downloaded),
            ("statements_transformed",       report.total_statements_transformed),
            ("statements_loaded",            report.total_statements_loaded),
            ("total_errors",                 len(report.all_errors)),
        ]

        for row_idx, (label, value) in enumerate(rows, start=1):
            label_cell       = ws.cell(row=row_idx, column=1, value=label)
            label_cell.font  = self.HEADER_FONT
            label_cell.fill  = self.HEADER_FILL
            label_cell.border = self.BORDER
            ws.column_dimensions["A"].width = 30

            value_cell        = ws.cell(row=row_idx, column=2, value=value)
            value_cell.border = self.BORDER
            ws.column_dimensions["B"].width = 25

            if isinstance(value, datetime):
                value_cell.number_format = 'YYYY-MM-DD HH:mm'

        # Per-company breakdown below the aggregate block
        ws.cell(row=len(rows) + 2, column=1, value="Per-company breakdown").font = self.HEADER_FONT

        company_headers = ["ticker", "downloaded", "transformed", "loaded", "duration_s", "errors"]
        company_widths  = [15, 14, 14, 14, 12, 8]
        self._write_header_row(ws, company_headers, company_widths, start_row=len(rows) + 3)

        for row_idx, cs in enumerate(report.company_stats, start=len(rows) + 4):
            self._write_data_row(ws, row_idx, [
                cs.ticker,
                cs.statements_downloaded,
                cs.statements_transformed,
                cs.statements_loaded,
                round(cs.duration_seconds, 2),
                len(cs.errors),
            ])

    def _write_errors_sheet(self, report: PipelineReport) -> None:
        ws = self.wb.create_sheet("errors")

        if not report.all_errors:
            ws.cell(row=1, column=1, value="No errors recorded.")
            return

        columns = [
            ("timestamp",            22),
            ("stage",                12),
            ("ticker",               12),
            ("error_type",           22),
            ("message",              60),
            ("document_description", 80),
        ]
        headers = [c[0] for c in columns]
        widths  = [c[1] for c in columns]
        self._write_header_row(ws, headers, widths)

        for row_idx, err in enumerate(report.all_errors, start=2):
            row_data = [
                err.timestamp,
                err.stage,
                err.ticker,
                err.error_type,
                err.message,
                err.document_description or "",
            ]
            self._write_data_row(ws, row_idx, row_data)

            # Highlight the entire error row in light red
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = self.ERROR_FILL

        ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Company sheets (unchanged)                                           #
    # ------------------------------------------------------------------ #

    def _write_company_metadata_sheet(self, company: Company) -> None:
        ws = self.wb.create_sheet("company_metadata")

        columns = [
            ("key", 25),
            ("value", 100)
        ]

        headers = [col[0] for col in columns]
        widths  = [col[1] for col in columns]
        self._write_header_row(ws, headers, widths)

        rows = [
            ("id", company.id),
            ("name", company.name),
            ("ticker", company.ticker),
            ("description", company.description),
            ("sector", company.sector),
            ("industry_group", company.industry_group),
            ("industry", company.industry),
            ("sub_industry", company.sub_industry),
            ("sub_industry_description", company.sub_industry_description)
        ]

        for row_idx, row_data in enumerate(rows, start=2):
            self._write_data_row(ws, row_idx, row_data)

    def _write_statement_metadata_sheet(self, company: Company) -> None:
        ws = self.wb.create_sheet("statement_metadata")

        columns = [
            ("document_id", 15),
            ("document_description", 100),
            ("submission_date", 20),
            ("period_end_date", 15),
            ("reporting_period", 15),
            ("financial_statements_type", 25),
            ("currency", 15),
            ("accounting_standards", 25),
            ("merger_or_demerger_in_process", 20),
            ("treasury_shares", 20),
            ("insolvency_proceedings", 20),
            ("public_offering_of_equity_instruments", 20),
            ("subscribed_shares", 20),
            ("share_price", 15),
            ("free_float_percentage", 15),
            ("number_of_employees", 20)
        ]

        headers = [col[0] for col in columns]
        widths  = [col[1] for col in columns]
        self._write_header_row(ws, headers, widths)

        for row_idx, stmt in enumerate(company.statements.values(), start=2):
            row_data = [
                stmt.document_id,
                stmt.document_description,
                stmt.submission_date,
                stmt.period_end_date,
                stmt.reporting_period,
                stmt.financial_statements_type,
                stmt.presentation_currency,
                stmt.accounting_standards_applied,
                stmt.merger_or_demerger_in_process,
                stmt.treasury_shares,
                stmt.insolvency_proceedings,
                stmt.public_offering_of_equity_instruments,
                stmt.subscribed_shares_at_period_end,
                stmt.equity_share_price_at_period_end,
                stmt.free_float_percentage,
                stmt.number_of_employees
            ]
            self._write_data_row(ws, row_idx, row_data)

        ws.freeze_panes = 'A2'

    def _write_statement_values_sheet(
        self,
        statements: Dict[int, CleanFinancialStatement]
    ) -> None:
        ws = self.wb.create_sheet("statement_values")

        columns = [
            ("document_id", 15),
            ("order", 12),
            ("concept", 40),
            ("value", 20)
        ]

        headers = [col[0] for col in columns]
        widths  = [col[1] for col in columns]
        self._write_header_row(ws, headers, widths)

        row_idx = 2
        for stmt in statements.values():
            for line in stmt.statement.values():
                self._write_data_row(ws, row_idx, [
                    stmt.document_id,
                    line.order,
                    line.label,
                    line.value
                ])
                row_idx += 1

        ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Shared helpers                                                       #
    # ------------------------------------------------------------------ #

    def _write_header_row(
        self,
        ws,
        headers: list[str],
        widths: list[int],
        start_row: int = 1
    ) -> None:
        for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value     = header
            cell.font      = self.HEADER_FONT
            cell.fill      = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = self.BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_data_row(self, ws, row_idx: int, data: list[Any]) -> None:
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if isinstance(value, datetime):
                cell.value         = value
                cell.number_format = 'YYYY-MM-DD HH:mm'
            elif isinstance(value, date):
                cell.value         = value
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, int):
                cell.value         = value
                cell.number_format = '#,##0'
            elif isinstance(value, float):
                cell.value         = value
                cell.number_format = '#,##0.00'
            else:
                cell.value = value

            cell.border = self.BORDER


# ------------------------------------------------------------------ #
# Convenience wrappers                                                #
# ------------------------------------------------------------------ #

def export_company_to_excel(company: Company, output_path: Path) -> Path:
    """Convenience wrapper around ExcelExporter for company data."""
    return ExcelExporter().export_company(company, output_path)


def export_report_to_excel(report: PipelineReport, output_path: Path) -> Path:
    """Convenience wrapper around ExcelExporter for pipeline reports."""
    return ExcelExporter().write_pipeline_report(report, output_path)