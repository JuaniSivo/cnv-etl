from pathlib import Path
from typing import Dict, Optional, get_args, get_origin
from dataclasses import dataclass, field, fields

from openpyxl import load_workbook

from cnv_etl.models.document import CleanFinancialStatement


@dataclass
class Company:
    id: int
    name: str
    ticker: str
    description: Optional[str] = None
    sector: Optional[str] = None
    industry_group: Optional[str] = None
    industry: Optional[str] = None
    sub_industry: Optional[str] = None
    sub_industry_description: Optional[str] = None
    statements: Dict[int, CleanFinancialStatement] = field(default_factory=dict)

    def add_statement(self, statement: CleanFinancialStatement) -> None:
        self.statements[statement.document_id] = statement


# ---------------------------------------------------------------------------
# Type-aware coercion
# ---------------------------------------------------------------------------

def _resolve_type(hint):
    """
    Unwrap Optional[X] to X.
    Returns the raw type annotation otherwise.
    """
    if get_origin(hint) is type(Optional[int]) or get_origin(hint) is type(None):
        pass

    args = get_args(hint)
    if args:
        # Optional[X] is Union[X, None] — return the first non-None arg
        non_none = [a for a in args if a is not type(None)]
        if non_none:
            return non_none[0]

    return hint


def _build_field_types(cls) -> dict[str, type]:
    """
    Return a mapping of field_name → resolved Python type for a dataclass,
    excluding fields with complex types (e.g. Dict, List).
    """
    hints = {f.name: f.type for f in fields(cls)}
    resolved = {}
    for name, hint in hints.items():
        # Skip already-evaluated annotations that are strings (forward refs)
        if isinstance(hint, str):
            continue
        t = _resolve_type(hint)
        # Only keep scalar types we can safely coerce to
        if t in (int, float, str, bool):
            resolved[name] = t
    return resolved


def _coerce(value, target_type: type):
    """
    Coerce a raw Excel cell value to target_type.
    Returns None for empty/null values.
    """
    if value is None:
        return None

    raw = str(value).strip()
    if raw == "" or raw == "None":
        return None

    try:
        if target_type is int:
            # Handle floats written as "30123456789.0" by Excel
            return int(float(raw))
        if target_type is float:
            return float(raw)
        if target_type is bool:
            return raw.lower() in ("true", "1", "yes", "si", "sí")
        return raw  # str
    except (ValueError, TypeError) as e:
        raise ValueError(f"Cannot coerce '{raw}' to {target_type.__name__}: {e}")


# ---------------------------------------------------------------------------
# Companies registry
# ---------------------------------------------------------------------------

class Companies:
    def __init__(self):
        self.by_ticker: Dict[str, Company] = dict()

    def add(self, company: Company) -> None:
        self.by_ticker[company.ticker] = company

    def get_by_ticker(self, ticker: str) -> Company:
        company = self.by_ticker.get(ticker)
        if company is not None:
            return company
        raise ValueError(
            f"Ticker '{ticker}' not found. Available tickers: {list(self.by_ticker.keys())}"
        )

    def load_from_excel(
        self,
        file_path: Path | str,
        sheet_name: str = "companies",
        column_mapping: Optional[Dict[str, str]] = None
    ) -> None:
        """
        Load companies from an Excel file.

        Parameters
        ----------
        file_path : Path | str
            Path to Excel file.
        sheet_name : str, default "companies"
            Name of the sheet containing company data.
        column_mapping : dict, optional
            Mapping from Excel column names to Company attribute names.
            If None, uses default mapping.

        Raises
        ------
        FileNotFoundError
            If the file doesn't exist.
        ValueError
            If the sheet doesn't exist or required columns are missing.
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if column_mapping is None:
            column_mapping = {
                "cuit":                   "id",
                "company_name":           "name",
                "ticker":                 "ticker",
                "company_description":    "description",
                "sector":                 "sector",
                "industry_group":         "industry_group",
                "industry":               "industry",
                "subindustry":            "sub_industry",
                "subindustry_definition": "sub_industry_description",
            }

        # Introspect Company field types once, before the loop
        field_types = _build_field_types(Company)

        wb = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        headers = [str(cell.value) for cell in ws[1]]

        missing = set(column_mapping.keys()) - set(headers)
        if missing:
            raise ValueError(
                f"Missing required columns in Excel: {missing}\n"
                f"Available columns: {headers}"
            )

        col_indices = {
            excel_col: headers.index(excel_col)
            for excel_col in column_mapping
            if excel_col in headers
        }

        companies_loaded = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            company_data = {}
            for excel_col, company_attr in column_mapping.items():
                if excel_col not in col_indices:
                    continue

                raw_value = row[col_indices[excel_col]]
                target_type = field_types.get(company_attr, str)

                try:
                    company_data[company_attr] = _coerce(raw_value, target_type)
                except ValueError as e:
                    print(f"⚠ Row {row_idx}, field '{company_attr}': {e}. Skipping field.")
                    company_data[company_attr] = None

            if not company_data.get("id") or not company_data.get("name"):
                print(f"⚠ Skipping row {row_idx}: missing id or name")
                continue

            try:
                company = Company(**company_data)
                self.add(company)
                companies_loaded += 1
            except Exception as e:
                print(f"⚠ Error loading company at row {row_idx}: {e}")

        wb.close()
        print(f"✓ Loaded {companies_loaded} companies from {file_path}")

    def __len__(self) -> int:
        return len(self.by_ticker)

    def __iter__(self):
        return iter(self.by_ticker.values())